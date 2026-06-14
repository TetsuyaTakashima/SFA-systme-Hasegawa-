#!/usr/bin/env python3
import csv
import html
import io
import re
import unicodedata
import urllib.parse
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "data" / "private-schools-k12.csv"
OUTPUT_CSV = ROOT / "data" / "private-schools-k12-contact-enriched.csv"
SUMMARY_MD = ROOT / "data" / "private-schools-k12-contact-summary.md"

USER_AGENT = "Mozilla/5.0 (compatible; airsetting-school-contact-enricher/1.0)"


TOKYO_SOURCES = [
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/03elementary_r80401_140",
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/06juniorhigh_r80401_140",
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/09high_zennichi_r80401_140",
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/12high_teiji_r80401_140",
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/15high_tushin_r61201",
    "https://www.seikatubunka.metro.tokyo.lg.jp/documents/d/seikatubunka/18special_r70401_140",
]

OSAKA_SOURCES = [
    "https://www.pref.osaka.lg.jp/o180160/shigaku/syoutyuukou/itiran-syou.html",
    "https://www.pref.osaka.lg.jp/o180160/shigaku/syoutyuukou/itiran-chugaku.html",
    "https://www.pref.osaka.lg.jp/o180160/shigaku/syoutyuukou/itiran-koukou.html",
    "https://www.pref.osaka.lg.jp/o180160/shigaku/syoutyuukou/itiran-chutou.html",
]

KANAGAWA_INDEX = "https://phsk.or.jp/school/"
CHIBA_INDEX = "https://chibashigaku.jp/hs/school/"
SAITAMA_PDFS = [
    "https://saitamashigaku.com/relays/download/6/26/6//?file=/files/libs/688/202504101629027726.pdf",
    "https://saitamashigaku.com/relays/download/6/26/5/688/?file=/files/libs/694/20250414133830563.pdf",
]
FUKUOKA_SEARCH = "https://www.f-sigaku.com/search/"
SHIZUOKA_XLSX = "https://www.shizuoka-shigaku.net/app/uploads/shizuoka-shigaku-member-school.xlsx"
HIROSHIMA_SOURCES = [
    "https://www.hiroshima-shigaku.com/find/category4/",
    "https://www.hiroshima-shigaku.com/find/category5/",
    "https://www.hiroshima-shigaku.com/find/category6/",
]
KYOTO_PDFS = [
    "https://www.pref.kyoto.jp/bunkyo/documents/03r8shougakkou.pdf",
    "https://www.pref.kyoto.jp/bunkyo/documents/04r8chuugakkou.pdf",
    "https://www.pref.kyoto.jp/bunkyo/documents/05r8koukou.pdf",
]
HOKKAIDO_PDFS = [
    "https://www.pref.hokkaido.lg.jp/fs/1/2/2/2/4/3/6/6/_/%E7%A7%81%E7%AB%8B%E5%B0%8F%E5%AD%A6%E6%A0%A1%E5%90%8D%E7%B0%BF(R7.9.1).pdf",
    "https://www.pref.hokkaido.lg.jp/fs/1/2/6/7/4/2/2/2/_/%E7%A7%81%E7%AB%8B%E4%B8%AD%E5%AD%A6%E6%A0%A1%E5%90%8D%E7%B0%BF(R7.9.1).pdf",
    "https://www.pref.hokkaido.lg.jp/fs/1/2/6/7/4/2/2/1/_/%E7%A7%81%E7%AB%8B%E9%AB%98%E7%AD%89%E5%AD%A6%E6%A0%A1%E5%90%8D%E7%B0%BF(R7.9.1).pdf",
    "https://www.pref.hokkaido.lg.jp/fs/1/2/2/2/4/3/6/8/_/%E7%A7%81%E7%AB%8B%E7%89%B9%E5%88%A5%E6%94%AF%E6%8F%B4%E5%AD%A6%E6%A0%A1%E5%90%8D%E7%B0%BF(R7.9.1).pdf",
]
HYOGO_SOURCES = [
    "https://web.pref.hyogo.lg.jp/kk35/pa15_000000003.html",
    "https://web.pref.hyogo.lg.jp/kk35/pa15_000000004.html",
    "https://web.pref.hyogo.lg.jp/kk35/pa15_000000005.html",
]
AICHI_PDFS = [
    "https://www.pref.aichi.jp/uploaded/life/588285_2703650_misc.pdf",
    "https://www.pref.aichi.jp/uploaded/life/588285_2703651_misc.pdf",
]
IBARAKI_SOURCES = [
    "https://kyoiku.pref.ibaraki.jp/gakko/private-schools/high-school/",
    "https://kyoiku.pref.ibaraki.jp/gakko/private-schools/elementary-junior-high-school/",
]
NAGANO_SOURCES = [
    ("https://www.pref.nagano.lg.jp/ken-manabi/kyoiku/gakko/shochu/shiritsusho/mebo.html", "小学校"),
    ("https://www.pref.nagano.lg.jp/ken-manabi/kyoiku/gakko/shochu/shiritsuchu/mebo.html", "中学校"),
    ("https://www.pref.nagano.lg.jp/ken-manabi/kyoiku/gakkoukou/koukou/koukou/mebo.html", "高等学校"),
]
OKAYAMA_HIGH_SCHOOL_LIST = "http://www.oka-shigaku.gr.jp/schoollist.html"
OKAYAMA_JUNIOR_SCHOOL_LIST = "http://www.oka-shigaku.gr.jp/schoollist-js.html"
MIYAGI_PDF = "https://www.pref.miyagi.jp/documents/6697/r7_meibo.pdf"
GIFU_PDFS = [
    ("https://www.pref.gifu.lg.jp/uploaded/attachment/494343.pdf", ["高等学校"]),
    ("https://www.pref.gifu.lg.jp/uploaded/attachment/494344.pdf", ["小学校", "中学校"]),
]
NAGASAKI_PDF = "https://www.pref.nagasaki.jp/uploads/2025/06/1750726260.pdf"
NARA_PDFS = [
    ("https://www.pref.nara.lg.jp/documents/14426/03_r8syougakkou.pdf", ["小学校"]),
    ("https://www.pref.nara.lg.jp/documents/14426/04_r8tyuugakkou.pdf", ["中学校", "中等教育学校"]),
    ("https://www.pref.nara.lg.jp/documents/14426/05_r8koutougakkou.pdf", ["高等学校", "中等教育学校"]),
]
KAGOSHIMA_PDFS = [
    ("https://www.pref.kagoshima.jp/ab04/kyoiku-bunka/school/shiritu/documents/5222_20250911165648-1.pdf", ["高等学校"]),
    ("https://www.pref.kagoshima.jp/ab04/kyoiku-bunka/school/shiritu/documents/5222_20250910152021-1.pdf", ["小学校", "中学校"]),
]
MIE_PDF = "https://www.pref.mie.lg.jp/common/content/001153698.pdf"
KANAGAWA_XLSX = "https://www.pref.kanagawa.jp/documents/22470/r8~meibo.xlsx"
KUMAMOTO_PDF = "https://www.pref.kumamoto.jp/uploaded/attachment/301756.pdf"
YAMAGUCHI_PDFS = [
    ("https://www.pref.yamaguchi.lg.jp/uploaded/attachment/238990.pdf", ["中学校"]),
    ("https://www.pref.yamaguchi.lg.jp/uploaded/attachment/238991.pdf", ["高等学校"]),
]
FUKUSHIMA_PDFS = [
    ("https://www.pref.fukushima.lg.jp/uploaded/attachment/717686.pdf", ["高等学校"]),
    ("https://www.pref.fukushima.lg.jp/uploaded/attachment/717687.pdf", ["小学校", "中学校"]),
]
TOCHIGI_PDF = "https://www.pref.tochigi.lg.jp/b05/education/gakkoukyouiku/ichiran/documents/r7_shiritsu_01.pdf"
NIIGATA_PDF = "https://www.pref.niigata.lg.jp/uploaded/attachment/489726.pdf"
MIYAZAKI_SCHOOL_LIST = "https://www.miyazaki-shigaku.jp/school/"

_TARGET_ROWS = None


def fetch_bytes(url):
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=30) as response:
        return response.read()


def fetch_text(url):
    raw = fetch_bytes(url)
    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "ignore")


def clean_text(value):
    text = html.unescape(re.sub(r"<.*?>", " ", str(value or ""), flags=re.S))
    return re.sub(r"\s+", " ", text).strip()


def normalize_name(value):
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("髙", "高").replace("ヶ", "ケ")
    text = re.sub(r"[\s・･/／()（）［］\[\]【】,、.。\-‐ー－]", "", text)
    for token in [
        "私立",
        "学校法人",
        "中学校高等学校",
        "中学高等学校",
        "中学校",
        "高等学校",
        "高等部",
        "中等部",
        "小学部",
        "初等部",
        "小学校",
        "高校",
        "中学",
        "中等教育学校",
        "義務教育学校",
        "特別支援学校",
        "高等専門学校",
    ]:
        text = text.replace(token, "")
    return text


def normalize_precise_name(value):
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("髙", "高").replace("ヶ", "ケ")
    text = re.sub(r"[\s・･/／()（）［］\[\]【】,、.。\-‐ー－]", "", text)
    for token in ["私立", "学校法人"]:
        text = text.replace(token, "")
    return text


def normalize_phone(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    leading_parenthesis = re.search(r"\((0\d{1,4})\)\s*(\d{1,4})[-－ー]?(\d{3,4})", text)
    if leading_parenthesis:
        return "-".join(leading_parenthesis.groups())
    parenthesized = re.search(r"(0\d{1,4})\((\d{1,4})\)\s*(\d{3,4})", text)
    if parenthesized:
        return "-".join(parenthesized.groups())
    match = re.search(r"0\d{1,4}[-－ー]\d{1,4}[-－ー]\d{3,4}", text)
    if not match:
        match = re.search(r"(?<!\d)0\d{9,10}(?!\d)", text)
    if not match:
        return ""
    phone = match.group(0).replace("－", "-").replace("ー", "-")
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("00") or len(digits) not in {10, 11}:
        return ""
    if "-" in phone:
        return phone
    if len(digits) == 10:
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}" if digits.startswith(("03", "06")) else f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return phone


def add_contact(contacts, prefecture, name, phone="", website="", source="", source_url=""):
    normalized_phone = normalize_phone(phone)
    website = str(website or "").strip()
    for variant in contact_name_variants(name):
        key = (prefecture, normalize_precise_name(variant))
        if not key[1]:
            continue
        current = contacts.get(key, {})
        contacts[key] = {
            "phone": current.get("phone") or normalized_phone,
            "website": current.get("website") or website,
            "source": current.get("source") or source,
            "source_url": current.get("source_url") or source_url,
        }


def contact_name_variants(name):
    text = clean_school_list_name(name)
    if not text:
        return []
    variants = {text}

    replacements = [
        ("高校", "高等学校"),
        ("中学・高等学校", "中学校・高等学校"),
        ("中学高等学校", "中学校・高等学校"),
        ("中・高等学校", "中学校・高等学校"),
        ("中等部・高等部", "中学校・高等学校"),
    ]
    for source, target in replacements:
        if source in text:
            variants.add(text.replace(source, target))

    candidate_texts = set(variants)
    for candidate in list(candidate_texts):
        for sep in ["・", " "]:
            if sep in candidate:
                parts = [part.strip() for part in candidate.split(sep) if part.strip()]
                variants.update(parts)

        match = re.match(r"(.+?)高等学校・中学校(.+)$", candidate)
        if match:
            base, suffix = match.groups()
            variants.add(f"{base}高等学校{suffix}")
            variants.add(f"{base}中学校{suffix}")

        match = re.match(r"(.+?)中学校・高等学校(?:・(.+?小学校))?$", candidate)
        if match:
            base, elementary = match.groups()
            variants.add(f"{base}中学校")
            variants.add(f"{base}高等学校")
            if elementary:
                if elementary.startswith(base):
                    variants.add(elementary)
                else:
                    variants.add(f"{base}{elementary}")

    return sorted(variants, key=len, reverse=True)


def load_tokyo_contacts(contacts):
    for url in TOKYO_SOURCES:
        text = fetch_bytes(url).decode("cp932")
        rows = list(csv.DictReader(io.StringIO(text)))
        for row in rows:
            add_contact(
                contacts,
                "東京都",
                row.get("学校名称", ""),
                phone=row.get("電話番号", ""),
                source="東京都 私立学校名簿",
                source_url=url,
            )


def table_rows(text):
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.S | re.I):
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_html, flags=re.S | re.I)
        values = [clean_text(cell) for cell in cells]
        if values:
            yield values


def load_osaka_contacts(contacts):
    for url in OSAKA_SOURCES:
        text = fetch_text(url)
        for values in table_rows(text):
            if len(values) < 4 or values[:4] == ["学校名", "郵便番号", "所在地", "電話番号"]:
                continue
            name, _postal, _address, phone = values[:4]
            if "学校" not in name and "学院" not in name:
                continue
            add_contact(
                contacts,
                "大阪府",
                name,
                phone=phone,
                source="大阪府 私立学校一覧",
                source_url=url,
            )


def load_kanagawa_contacts(contacts):
    index_text = fetch_text(KANAGAWA_INDEX)
    school_links = []
    seen = set()
    for match in re.finditer(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', index_text, flags=re.S | re.I):
        label = clean_text(match.group(2))
        href = urllib.parse.urljoin(KANAGAWA_INDEX, html.unescape(match.group(1)))
        if not re.search(r"/school/\d+/?$", href) or href in seen:
            continue
        seen.add(href)
        school_links.append((label, href))

    for label, url in school_links:
        try:
            text = fetch_text(url)
        except Exception:
            continue
        rendered = clean_text(text)
        phone_match = re.search(r"TEL\s*([0-9０-９][0-9０-９\-－ー〜～\s]+)", rendered)
        phone = normalize_phone(phone_match.group(1) if phone_match else "")
        website_match = re.search(r"学校HP\s*(https?://\S+)", rendered)
        website = website_match.group(1).rstrip("。") if website_match else ""
        add_contact(
            contacts,
            "神奈川県",
            label,
            phone=phone,
            website=website,
            source="神奈川県私立中学高等学校協会 学校案内",
            source_url=url,
        )


def load_chiba_contacts(contacts):
    index_text = fetch_text(CHIBA_INDEX)
    school_links = []
    seen = set()
    for match in re.finditer(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', index_text, flags=re.S | re.I):
        href = urllib.parse.urljoin(CHIBA_INDEX, html.unescape(match.group(1)))
        if not re.search(r"/hs/school/\d+/?$", href) or href in seen:
            continue
        seen.add(href)
        school_links.append(href)

    for url in school_links:
        try:
            text = fetch_text(url)
        except Exception:
            continue
        heading = re.search(r"<h1[^>]*>(.*?)</h1>", text, flags=re.S | re.I)
        name = clean_text(heading.group(1)) if heading else ""
        rendered = clean_text(text)
        phone_match = re.search(r"電話番号\s*([0-9０-９][0-9０-９\-－ー\s]+)", rendered)
        website_match = re.search(r"ホームページ\s*(https?://\S+)", rendered)
        add_contact(
            contacts,
            "千葉県",
            name,
            phone=phone_match.group(1) if phone_match else "",
            website=website_match.group(1).rstrip("。") if website_match else "",
            source="千葉県私立中学高等学校協会 学校紹介",
            source_url=url,
        )


def extract_pdf_text(url):
    try:
        from pypdf import PdfReader
    except Exception as error:
        raise RuntimeError("pypdf is required to read PDF school lists") from error

    raw = fetch_bytes(url)
    reader = PdfReader(io.BytesIO(raw))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def extract_pdf_text_layout(url):
    try:
        from pypdf import PdfReader
    except Exception as error:
        raise RuntimeError("pypdf is required to read PDF school lists") from error

    raw = fetch_bytes(url)
    reader = PdfReader(io.BytesIO(raw))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text(extraction_mode="layout") or "")
        except Exception:
            pages.append(page.extract_text() or "")
    return "\n".join(pages)


def load_saitama_contacts(contacts):
    for url in SAITAMA_PDFS:
        text = extract_pdf_text(url)
        pending = ""
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            if not line:
                pending = ""
                continue
            if not normalize_phone(line):
                if "学校" in line or "学院" in line or "大学系属" in line:
                    pending = line
                continue

            combined = f"{pending} {line}".strip() if pending else line
            pending = ""
            postal_match = re.search(r"\d{3}-\d{4}", combined)
            if not postal_match:
                continue
            name = combined[: postal_match.start()].strip()
            name = re.sub(r"（.*?）|\(.*?\)", "", name).strip()
            if not name or ("学校" not in name and "学院" not in name):
                continue
            add_contact(
                contacts,
                "埼玉県",
                name,
                phone=normalize_phone(combined),
                source="埼玉県私立中学高等学校協会 学校名簿PDF",
                source_url=url,
            )


def load_fukuoka_contacts(contacts):
    text = fetch_text(FUKUOKA_SEARCH)
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.S | re.I):
        name_match = re.search(r"<h6[^>]*>(.*?)</h6>", row_html, flags=re.S | re.I)
        if not name_match:
            continue
        name = clean_text(name_match.group(1))
        phone_match = re.search(r'<dl[^>]*class="[^"]*telBox[^"]*"[^>]*>.*?<dd[^>]*>(.*?)</dd>', row_html, flags=re.S | re.I)
        web_match = re.search(r'<p[^>]*class="web"[^>]*>\s*<a[^>]+href="([^"]+)"', row_html, flags=re.S | re.I)
        add_contact(
            contacts,
            "福岡県",
            name,
            phone=clean_text(phone_match.group(1)) if phone_match else "",
            website=html.unescape(web_match.group(1)).strip() if web_match else "",
            source="福岡県私学協会 学校検索",
            source_url=FUKUOKA_SEARCH,
        )


def read_xlsx_rows(raw):
    namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("a:si", namespace):
                shared_strings.append("".join(text.text or "" for text in item.findall(".//a:t", namespace)))

        sheet_names = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
        if not sheet_names:
            return []
        root = ET.fromstring(archive.read(sheet_names[0]))
        rows = []
        for row in root.findall(".//a:row", namespace):
            values = []
            for cell in row.findall("a:c", namespace):
                value = cell.find("a:v", namespace)
                text = value.text if value is not None else ""
                if cell.attrib.get("t") == "s" and text:
                    text = shared_strings[int(text)]
                values.append(text)
            rows.append(values)
        return rows


def load_shizuoka_contacts(contacts):
    rows = read_xlsx_rows(fetch_bytes(SHIZUOKA_XLSX))
    for row in rows:
        if len(row) < 11:
            continue
        name = row[2]
        phone = row[10]
        if not name or not normalize_phone(phone):
            continue
        add_contact(
            contacts,
            "静岡県",
            name,
            phone=phone,
            source="静岡県私学協会 加盟校一覧XLSX",
            source_url=SHIZUOKA_XLSX,
        )


def load_hiroshima_contacts(contacts):
    for url in HIROSHIMA_SOURCES:
        text = fetch_text(url)
        for values in table_rows(text):
            if len(values) < 4 or values[:4] == ["学校名", "住所", "電話番号", "WEBサイト"]:
                continue
            name, _address, phone, _website_label = values[:4]
            if "学校" not in name and "学院" not in name and "高等" not in name:
                continue
            add_contact(
                contacts,
                "広島県",
                name,
                phone=phone,
                source="広島県私立中学高等学校協会 学校一覧",
                source_url=url,
            )


def pdf_contact_rows(url):
    text = extract_pdf_text(url)
    pending = ""
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            continue
        combined = f"{pending} {line}".strip() if pending else line
        if re.search(r"\d{3}-\d{4}", combined):
            if normalize_phone(combined) or re.search(r"\(\d{2,4}\)[-－ー]?\d{3,4}", unicodedata.normalize("NFKC", combined)):
                yield combined
                pending = ""
            else:
                pending = combined
        elif pending:
            if normalize_phone(combined) or re.search(r"\(\d{2,4}\)[-－ー]?\d{3,4}", unicodedata.normalize("NFKC", combined)):
                yield combined
                pending = ""
            else:
                pending = combined


def pdf_row_name(row_text):
    postal_match = re.search(r"\d{3}-\d{4}", row_text)
    if not postal_match:
        return ""
    name = row_text[: postal_match.start()].strip()
    name = re.sub(r"^\d+\s*", "", name).strip()
    name = re.sub(r"^[\-－ー\.\s]+", "", name).strip()
    if not name or name in {"学校名", "番号", "学 校 名"}:
        return ""
    return name


def kyoto_phone(row_text):
    phone = normalize_phone(row_text)
    if phone:
        return phone
    match = re.search(r"\((\d{2,4})\)[-－ー]?(\d{3,4})", unicodedata.normalize("NFKC", row_text))
    if not match:
        return ""
    return f"075-{match.group(1)}-{match.group(2)}"


def append_school_level(name, level):
    if level in name or "初等部" in name or "中等部" in name or "高等部" in name:
        return name
    return f"{name}{level}"


def clean_school_list_name(name):
    text = clean_text(name)
    text = re.sub(r"【.*?】|\(.*?\)|（.*?）", "", text)
    return re.sub(r"\s+", " ", text).strip()


def target_rows():
    global _TARGET_ROWS
    if _TARGET_ROWS is None:
        with INPUT_CSV.open(encoding="utf-8-sig", newline="") as source:
            _TARGET_ROWS = list(csv.DictReader(source))
    return _TARGET_ROWS


def target_rows_for(prefecture, school_types=None):
    rows = [row for row in target_rows() if row.get("都道府県") == prefecture]
    if not school_types:
        return rows
    return [row for row in rows if any(term in row.get("種別", "") or term in row.get("施設名", "") for term in school_types)]


def compact_search_text(value):
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("髙", "高").replace("ヶ", "ケ")
    chars = []
    mapping = []
    for index, char in enumerate(text):
        if re.match(r"[\s・･/／()（）［］\[\]【】,、.。\-‐ー－]", char):
            continue
        chars.append(char)
        mapping.append(index)
    return "".join(chars), mapping, text


def add_contacts_from_text_by_targets(contacts, prefecture, text, source, source_url, school_types=None, window=900):
    compact, mapping, normalized_text = compact_search_text(text)
    for row in target_rows_for(prefecture, school_types):
        name = row.get("施設名", "")
        keys = [normalize_precise_name(name), normalize_name(name)]
        seen = set()
        for key in keys:
            if not key or key in seen or len(key) < 4:
                continue
            seen.add(key)
            index = compact.find(key)
            if index == -1:
                continue
            original_index = mapping[index]
            snippet = normalized_text[original_index : original_index + window]
            postal_match = re.search(r"\d{3}-\d{4}", snippet)
            if postal_match:
                snippet = snippet[postal_match.end() :]
            phone = normalize_phone(snippet)
            if not phone:
                continue
            add_contact(
                contacts,
                prefecture,
                name,
                phone=phone,
                source=source,
                source_url=source_url,
            )
            break


def load_kyoto_contacts(contacts):
    for url in KYOTO_PDFS:
        level = "小学校" if "shougakkou" in url else "中学校" if "chuugakkou" in url else "高等学校"
        for row_text in pdf_contact_rows(url):
            name = pdf_row_name(row_text)
            phone = kyoto_phone(row_text)
            if not name or not phone:
                continue
            add_contact(
                contacts,
                "京都府",
                append_school_level(name, level),
                phone=phone,
                source="京都府 私立学校名簿PDF",
                source_url=url,
            )


def load_hokkaido_contacts(contacts):
    for url in HOKKAIDO_PDFS:
        text = extract_pdf_text(url)
        current_area = ""
        pending = ""
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            if not line:
                continue
            area_line = unicodedata.normalize("NFKC", line)
            area_match = re.search(r"【[^】]+】.*?(0\d{2,4})", area_line)
            if area_match:
                current_area = area_match.group(1)

            combined = f"{pending} {line}".strip() if pending else line
            if re.search(r"\d{3}-\d{4}", combined):
                pending = combined
                if not (normalize_phone(combined) or re.search(r"\d{2,4}-\d{4}$", unicodedata.normalize("NFKC", combined))):
                    continue
            elif pending:
                pending = combined
            else:
                continue

            name = pdf_row_name(combined)
            if not name:
                pending = ""
                continue
            phone = normalize_phone(combined)
            if not phone:
                suffix_match = re.search(r"(\d{2,4}-\d{4})$", unicodedata.normalize("NFKC", combined))
                if suffix_match and current_area:
                    phone = f"{current_area}-{suffix_match.group(1)}"
            if phone:
                add_contact(
                    contacts,
                    "北海道",
                    name,
                    phone=phone,
                    source="北海道 私立学校名簿PDF",
                    source_url=url,
                )
                pending = ""


def load_hyogo_contacts(contacts):
    for url in HYOGO_SOURCES:
        text = fetch_text(url)
        for values in table_rows(text):
            if len(values) < 4 or values[:4] == ["学校名", "郵便番号", "住所", "電話"]:
                continue
            name, _postal, _address, phone = values[:4]
            if not normalize_phone(phone):
                continue
            add_contact(
                contacts,
                "兵庫県",
                name,
                phone=phone,
                source="兵庫県 私立学校一覧",
                source_url=url,
            )


def load_aichi_contacts(contacts):
    for url in AICHI_PDFS:
        for row_text in pdf_contact_rows(url):
            name = pdf_row_name(row_text)
            phone = normalize_phone(row_text)
            if not name or not phone:
                continue
            add_contact(
                contacts,
                "愛知県",
                name,
                phone=phone,
                source="愛知県 私立学校名簿PDF",
                source_url=url,
            )


def load_ibaraki_contacts(contacts):
    for url in IBARAKI_SOURCES:
        text = fetch_text(url)
        for item_html in re.findall(r'<div[^>]+class="[^"]*private-school-item[^"]*"[^>]*>(.*?)<!-- cell -->', text, flags=re.S | re.I):
            heading = re.search(r"<h5[^>]*>(.*?)</h5>", item_html, flags=re.S | re.I)
            phone_match = re.search(r"<th[^>]*>\s*電話番号\s*</th>\s*<td[^>]*>(.*?)</td>", item_html, flags=re.S | re.I)
            if not heading or not phone_match:
                continue
            name = clean_school_list_name(heading.group(1))
            phone = clean_text(phone_match.group(1))
            if not normalize_phone(phone):
                continue
            add_contact(
                contacts,
                "茨城県",
                name,
                phone=phone,
                source="茨城県 私立学校一覧",
                source_url=url,
            )


def load_nagano_contacts(contacts):
    for url, level in NAGANO_SOURCES:
        text = fetch_text(url)
        for values in table_rows(text):
            if len(values) < 6 or values[0].endswith("名"):
                continue
            name = clean_school_list_name(values[0])
            phone = values[-1]
            if not normalize_phone(phone):
                continue
            add_contact(
                contacts,
                "長野県",
                append_school_level(name, level),
                phone=phone,
                source="長野県 私立学校名簿",
                source_url=url,
            )


def load_okayama_contacts(contacts):
    for url, default_level in [(OKAYAMA_HIGH_SCHOOL_LIST, "高等学校"), (OKAYAMA_JUNIOR_SCHOOL_LIST, "")]:
        text = fetch_text(url)
        junior_header_count = 0
        junior_level = "中等教育学校"
        for values in table_rows(text):
            if len(values) < 5:
                continue
            if values[0] == "学校名":
                if url == OKAYAMA_JUNIOR_SCHOOL_LIST:
                    junior_header_count += 1
                    junior_level = "中等教育学校" if junior_header_count == 1 else "中学校"
                continue
            name = clean_school_list_name(values[0])
            phone = values[-1]
            if not normalize_phone(phone):
                continue
            level = default_level or junior_level
            add_contact(
                contacts,
                "岡山県",
                append_school_level(name, level),
                phone=phone,
                source="岡山県私学協会 学校一覧",
                source_url=url,
            )


def load_miyagi_contacts(contacts):
    text = extract_pdf_text_layout(MIYAGI_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "宮城県",
        text,
        "宮城県 私立学校名簿PDF",
        MIYAGI_PDF,
        ["小学校", "中学校", "高等学校", "特別支援学校"],
    )


def load_gifu_contacts(contacts):
    for url, school_types in GIFU_PDFS:
        text = extract_pdf_text_layout(url)
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            if not any(school_type in line for school_type in school_types):
                continue
            if "学校法人" not in line:
                continue
            name_match = re.match(r"^\d+\s+(.+?学校)\s+学校法人", line)
            phone_match = re.search(r"[（(]\d{3}-\d{4}[）)]\s*(0\d{1,4}-\d{1,4}-\d{3,4})", line)
            if not name_match or not phone_match:
                continue
            add_contact(
                contacts,
                "岐阜県",
                name_match.group(1),
                phone=phone_match.group(1),
                source="岐阜県 私立学校名簿PDF",
                source_url=url,
            )


def load_nagasaki_contacts(contacts):
    text = extract_pdf_text(NAGASAKI_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "長崎県",
        text,
        "長崎県 私立小・中・高等学校一覧PDF",
        NAGASAKI_PDF,
        ["小学校", "中学校", "高等学校"],
        window=700,
    )


def load_nara_contacts(contacts):
    for url, school_types in NARA_PDFS:
        text = extract_pdf_text(url)
        add_contacts_from_text_by_targets(
            contacts,
            "奈良県",
            text,
            "奈良県 私立学校名簿PDF",
            url,
            school_types,
            window=450,
        )


def load_kagoshima_contacts(contacts):
    for url, school_types in KAGOSHIMA_PDFS:
        text = extract_pdf_text(url)
        add_contacts_from_text_by_targets(
            contacts,
            "鹿児島県",
            text,
            "鹿児島県 私立学校名簿PDF",
            url,
            school_types,
            window=700,
        )


def load_mie_contacts(contacts):
    text = extract_pdf_text(MIE_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "三重県",
        text,
        "三重県 学校名簿PDF",
        MIE_PDF,
        ["小学校", "中学校", "高等学校", "中等教育学校", "特別支援学校", "高等専門学校"],
        window=700,
    )


def load_kanagawa_official_contacts(contacts):
    try:
        import openpyxl
    except Exception as error:
        raise RuntimeError("openpyxl is required to read Kanagawa school list") from error

    workbook = openpyxl.load_workbook(io.BytesIO(fetch_bytes(KANAGAWA_XLSX)), data_only=True)
    for sheet_name in ["小学校", "中学校", "高等学校", "中等教育学校", "特別支援学校"]:
        worksheet = workbook[sheet_name]
        header = None
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [str(value).replace("\n", "") if value is not None else "" for value in row]
            if "学校名" in values and "電話番号" in values:
                header = (row_index, values.index("学校名"), values.index("電話番号"))
                break
        if not header:
            continue
        _header_row, name_col, phone_col = header
        for row in worksheet.iter_rows(min_row=header[0] + 1, values_only=True):
            name = clean_school_list_name(row[name_col] if len(row) > name_col else "")
            phone = row[phone_col] if len(row) > phone_col else ""
            if not name or not normalize_phone(phone):
                continue
            add_contact(
                contacts,
                "神奈川県",
                name,
                phone=str(phone),
                source="神奈川県 私立学校名簿XLSX",
                source_url=KANAGAWA_XLSX,
            )


def load_kumamoto_contacts(contacts):
    text = extract_pdf_text(KUMAMOTO_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "熊本県",
        text,
        "熊本県 私立学校名簿PDF",
        KUMAMOTO_PDF,
        ["小学校", "中学校", "高等学校"],
        window=700,
    )


def load_yamaguchi_contacts(contacts):
    for url, school_types in YAMAGUCHI_PDFS:
        text = extract_pdf_text(url)
        add_contacts_from_text_by_targets(
            contacts,
            "山口県",
            text,
            "山口県 私立学校一覧PDF",
            url,
            school_types,
            window=600,
        )


def load_fukushima_contacts(contacts):
    for url, school_types in FUKUSHIMA_PDFS:
        text = extract_pdf_text(url)
        add_contacts_from_text_by_targets(
            contacts,
            "福島県",
            text,
            "福島県 私立学校一覧PDF",
            url,
            school_types,
            window=700,
        )


def load_tochigi_contacts(contacts):
    text = extract_pdf_text(TOCHIGI_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "栃木県",
        text,
        "栃木県 私立学校一覧PDF",
        TOCHIGI_PDF,
        ["小学校", "中学校", "高等学校", "中等教育学校"],
        window=700,
    )


def load_niigata_contacts(contacts):
    text = extract_pdf_text(NIIGATA_PDF)
    add_contacts_from_text_by_targets(
        contacts,
        "新潟県",
        text,
        "新潟県 私立中学・高等学校一覧PDF",
        NIIGATA_PDF,
        ["中学校", "高等学校"],
        window=700,
    )


def load_miyazaki_contacts(contacts):
    text = fetch_text(MIYAZAKI_SCHOOL_LIST)
    for item_html in re.findall(r'<div[^>]+class="[^"]*listBox[^"]*"[^>]*>(.*?)</div>\s*</div>\s*</div>', text, flags=re.S | re.I):
        name_match = re.search(r'<div[^>]+class="[^"]*itemList name[^"]*"[^>]*>\s*<span>(.*?)</span>', item_html, flags=re.S | re.I)
        if not name_match:
            continue
        phone_match = re.search(r'<div[^>]+class="[^"]*itemTtl[^"]*"[^>]*>\s*TEL\s*</div>\s*<div[^>]+class="[^"]*item[^"]*"[^>]*>(.*?)</div>', item_html, flags=re.S | re.I)
        if not phone_match:
            continue
        name = clean_school_list_name(name_match.group(1))
        phone = clean_text(phone_match.group(1))
        if not normalize_phone(phone):
            continue
        add_contact(
            contacts,
            "宮崎県",
            name,
            phone=phone,
            source="宮崎県私立中学高等学校協会 学校一覧",
            source_url=MIYAZAKI_SCHOOL_LIST,
        )


def find_contact(contacts, row):
    prefecture = row.get("都道府県", "")
    direct_key = (prefecture, normalize_precise_name(row.get("施設名", "")))
    if direct_key in contacts:
        return contacts[direct_key]

    name_key = normalize_name(row.get("施設名", ""))
    if not name_key:
        return None
    candidates = [(key, value) for key, value in contacts.items() if key[0] == prefecture]
    matches = []
    for (_pref, source_key), value in candidates:
        source_base = normalize_name(source_key)
        if len(name_key) >= 4 and source_base and (name_key == source_base or name_key in source_base or source_base in name_key):
            matches.append(value)
    phones = {match.get("phone") for match in matches if match.get("phone")}
    if len(phones) == 1:
        return matches[0]
    return None


def main():
    contacts = {}
    loaders = [
        ("東京都", load_tokyo_contacts),
        ("大阪府", load_osaka_contacts),
        ("神奈川県公式", load_kanagawa_official_contacts),
        ("神奈川県", load_kanagawa_contacts),
        ("千葉県", load_chiba_contacts),
        ("埼玉県", load_saitama_contacts),
        ("福岡県", load_fukuoka_contacts),
        ("静岡県", load_shizuoka_contacts),
        ("広島県", load_hiroshima_contacts),
        ("京都府", load_kyoto_contacts),
        ("北海道", load_hokkaido_contacts),
        ("兵庫県", load_hyogo_contacts),
        ("愛知県", load_aichi_contacts),
        ("茨城県", load_ibaraki_contacts),
        ("長野県", load_nagano_contacts),
        ("岡山県", load_okayama_contacts),
        ("宮城県", load_miyagi_contacts),
        ("岐阜県", load_gifu_contacts),
        ("長崎県", load_nagasaki_contacts),
        ("奈良県", load_nara_contacts),
        ("鹿児島県", load_kagoshima_contacts),
        ("三重県", load_mie_contacts),
        ("熊本県", load_kumamoto_contacts),
        ("山口県", load_yamaguchi_contacts),
        ("福島県", load_fukushima_contacts),
        ("栃木県", load_tochigi_contacts),
        ("新潟県", load_niigata_contacts),
        ("宮崎県", load_miyazaki_contacts),
    ]
    loader_results = []
    for name, loader in loaders:
        before = len(contacts)
        try:
            loader(contacts)
            loader_results.append((name, len(contacts) - before, "OK"))
        except Exception as error:
            loader_results.append((name, 0, f"ERROR: {error}"))

    with INPUT_CSV.open(encoding="utf-8-sig", newline="") as source:
        rows = list(csv.DictReader(source))

    output_fields = list(rows[0].keys()) + ["補完状況", "補完元", "補完元URL"]
    matched = 0
    phone_count = 0
    website_count = 0
    enriched_rows = []
    for row in rows:
        next_row = dict(row)
        contact = find_contact(contacts, row)
        if contact:
            matched += 1
            if not next_row.get("電話番号") and contact.get("phone"):
                next_row["電話番号"] = contact["phone"]
            if not next_row.get("Webサイト") and contact.get("website"):
                next_row["Webサイト"] = contact["website"]
            if next_row.get("電話番号"):
                phone_count += 1
            if next_row.get("Webサイト"):
                website_count += 1
            next_row["補完状況"] = "補完済み"
            next_row["補完元"] = contact.get("source", "")
            next_row["補完元URL"] = contact.get("source_url", "")
        else:
            next_row["補完状況"] = "未補完"
            next_row["補完元"] = ""
            next_row["補完元URL"] = ""
        enriched_rows.append(next_row)

    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=output_fields)
        writer.writeheader()
        writer.writerows(enriched_rows)

    by_pref = {}
    for row in enriched_rows:
        pref = row.get("都道府県", "")
        stats = by_pref.setdefault(pref, {"total": 0, "matched": 0, "phone": 0, "website": 0})
        stats["total"] += 1
        stats["matched"] += row["補完状況"] == "補完済み"
        stats["phone"] += bool(row.get("電話番号"))
        stats["website"] += bool(row.get("Webサイト"))

    lines = [
        "# 私立小中高等学校 連絡先補完",
        "",
        f"- 対象: {len(rows):,}件",
        f"- 補完済み: {matched:,}件",
        f"- 電話番号あり: {phone_count:,}件",
        f"- Webサイトあり: {website_count:,}件",
        "",
        "## 取得元",
        "",
    ]
    lines += [f"- {name}: {count:,}件取得 ({status})" for name, count, status in loader_results]
    lines += ["", "## 都道府県別", ""]
    for pref, stats in sorted(by_pref.items()):
        if stats["matched"] or pref in {"東京都", "神奈川県", "大阪府"}:
            lines.append(
                f"- {pref}: 対象{stats['total']:,}件 / 補完{stats['matched']:,}件 / 電話{stats['phone']:,}件 / Web{stats['website']:,}件"
            )
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"contacts={len(contacts)} matched={matched} phones={phone_count} websites={website_count}")
    print(OUTPUT_CSV)
    print(SUMMARY_MD)


if __name__ == "__main__":
    main()
